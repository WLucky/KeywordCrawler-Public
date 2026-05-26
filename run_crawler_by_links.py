#!/usr/bin/env python3
# -*- coding: utf-8 -*-
import re
import sys
import os
import json
import time
from pathlib import Path
from datetime import datetime

try:
    import openpyxl
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))


class Unbuffered:
    def __init__(self, stream):
        self.stream = stream

    def write(self, data):
        self.stream.write(data)
        self.stream.flush()

    def writelines(self, datas):
        self.stream.writelines(datas)
        self.stream.flush()

    def __getattr__(self, attr):
        return getattr(self.stream, attr)


sys.stdout = Unbuffered(sys.stdout)


def update_config_for_links(platform, video_links_str, max_comments, enable_sub_comments, enable_video_download):
    """更新配置文件用于视频链接爬取"""
    
    video_list = [link.strip() for link in video_links_str.split(';') if link.strip()]
    
    if 'dy' in platform.lower():
        dy_config_file = 'config/dy_config.py'
        with open(dy_config_file, 'r', encoding='utf-8') as f:
            dy_content = f.read()

        new_list_str = ',\n    '.join([f'"{link}"' for link in video_list])

        pattern = r'DY_SPECIFIED_ID_LIST = \[.*?\]'
        replacement = f'DY_SPECIFIED_ID_LIST = [\n    {new_list_str},\n]'
        dy_content = re.sub(pattern, replacement, dy_content, flags=re.DOTALL)

        dy_content = re.sub(r'ENABLE_DY_VIDEO_DOWNLOAD = \w+', f'ENABLE_DY_VIDEO_DOWNLOAD = {enable_video_download}', dy_content)

        with open(dy_config_file, 'w', encoding='utf-8') as f:
            f.write(dy_content)

        print(f'已更新DY_SPECIFIED_ID_LIST，共 {len(video_list)} 个视频链接')

    if 'bili' in platform.lower():
        bili_config_file = 'config/bilibili_config.py'
        with open(bili_config_file, 'r', encoding='utf-8') as f:
            bili_content = f.read()

        new_list_str = ',\n    '.join([f'"{link}"' for link in video_list])

        pattern = r'(BILI_SPECIFIED_ID_LIST = \[)\s*[\s\S]*?(\])'
        replacement = rf'\1\n    {new_list_str},\n\2'
        bili_content = re.sub(pattern, replacement, bili_content)

        with open(bili_config_file, 'w', encoding='utf-8') as f:
            f.write(bili_content)

        print(f'已更新BILI_SPECIFIED_ID_LIST，共 {len(video_list)} 个视频链接')

    base_config_file = 'config/base_config.py'
    with open(base_config_file, 'r', encoding='utf-8') as f:
        base_content = f.read()

    base_content = re.sub(r'PLATFORM = ".*"', f'PLATFORM = "{platform.split(",")[0]}"', base_content)

    base_content = re.sub(
        r'(CRAWLER_TYPE = \(\s*)"search"',
        r'\1"detail"',
        base_content
    )

    base_content = re.sub(r'CRAWLER_MAX_NOTES_COUNT = \d+', f'CRAWLER_MAX_NOTES_COUNT = {max_comments}', base_content)
    base_content = re.sub(r'CRAWLER_MAX_COMMENTS_COUNT_SINGLENOTES = \d+', f'CRAWLER_MAX_COMMENTS_COUNT_SINGLENOTES = {max_comments}', base_content)
    base_content = re.sub(r'ENABLE_GET_SUB_COMMENTS = \w+', f'ENABLE_GET_SUB_COMMENTS = {enable_sub_comments}', base_content)
    base_content = re.sub(r'ENABLE_GET_MEIDAS = \w+', f'ENABLE_GET_MEIDAS = {enable_video_download}', base_content)

    with open(base_config_file, 'w', encoding='utf-8') as f:
        f.write(base_content)

    print(f'已更新配置: PLATFORM={platform}, CRAWLER_TYPE=detail')
    print(f'评论数量: {max_comments}, 二级评论: {enable_sub_comments}')
    print(f'视频下载: {enable_video_download}')


def get_platform_name(platform):
    platform_names = {
        'dy': '抖音',
        'douyin': '抖音',
        'bili': '哔哩哔哩',
        'bilibili': '哔哩哔哩',
        'xhs': '小红书',
        'xiaohongshu': '小红书',
        'ks': '快手',
        'kuaishou': '快手',
        'wb': '微博',
        'weibo': '微博',
        'zhihu': '知乎',
        'tieba': '贴吧',
        'tavily': 'Tavily'
    }
    return platform_names.get(platform.lower(), platform)


def extract_video_links(item, video_links, platform):
    url = None
    title = None

    url_fields = [
        'video_download_url',
        'video_url',
        'aweme_url',
        'video_play_url',
        'note_url',
        'content_url',
        'download_url'
    ]

    title_fields = [
        'title',
        'desc',
        'aweme_id',
        'video_id',
        'note_id',
        'content_id'
    ]

    for field in url_fields:
        if field in item and item[field]:
            url = item[field]
            break

    for field in title_fields:
        if field in item and item[field]:
            title = str(item[field])
            break

    if url:
        if not title:
            title = f"{get_platform_name(platform)}视频"
        video_links.append((title, url))


def generate_video_download_markdown(platforms, keywords):
    data_dir = Path('data')
    data_dir.mkdir(exist_ok=True)
    markdown_lines = []

    timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    markdown_lines.append(f"# 视频下载链接汇总")
    markdown_lines.append(f"")
    markdown_lines.append(f"**搜索关键词**: {keywords}")
    markdown_lines.append(f"**生成时间**: {timestamp}")
    markdown_lines.append(f"")

    for platform in platforms:
        markdown_lines.append(f"## {get_platform_name(platform)}")
        markdown_lines.append(f"")

        content_files = []
        platform_name = get_platform_name(platform)
        platform_full_names = {
            'dy': 'douyin',
            'bili': 'bilibili',
            'xhs': 'xiaohongshu',
            'ks': 'kuaishou',
            'wb': 'weibo'
        }

        csv_dir = data_dir / 'csv'
        if csv_dir.exists():
            for ext in ['csv']:
                content_files.extend(list(csv_dir.glob(f'*{platform}*contents*.{ext}')))
                content_files.extend(list(csv_dir.glob(f'*{platform_name}*contents*.{ext}')))
                if platform in platform_full_names:
                    content_files.extend(list(csv_dir.glob(f'*{platform_full_names[platform]}*contents*.{ext}')))

        xlsx_dir = data_dir / 'xlsx'
        if xlsx_dir.exists():
            for ext in ['xlsx']:
                content_files.extend(list(xlsx_dir.glob(f'{platform}_*.{ext}')))
                content_files.extend(list(xlsx_dir.glob(f'{platform_name}_*.{ext}')))
                if platform in platform_full_names:
                    content_files.extend(list(xlsx_dir.glob(f'{platform_full_names[platform]}_*.{ext}')))

        platform_dir = data_dir / platform
        if platform_dir.exists():
            for ext in ['json', 'jsonl', 'csv', 'xlsx']:
                content_files.extend(list(platform_dir.glob(f'**/*contents*.{ext}')))
                content_files.extend(list(platform_dir.glob(f'**/*video*.{ext}')))
                content_files.extend(list(platform_dir.glob(f'**/*{platform}*.{ext}')))
                content_files.extend(list(platform_dir.glob(f'**/*search*.{ext}')))
                content_files.extend(list(platform_dir.glob(f'*{platform}*.{ext}')))
                content_files.extend(list(platform_dir.glob(f'*search*.{ext}')))

        for ext in ['json', 'jsonl', 'xlsx']:
            content_files.extend(list(data_dir.glob(f'*{platform}*contents*.{ext}')))
            content_files.extend(list(data_dir.glob(f'*{platform}*video*.{ext}')))
            content_files.extend(list(data_dir.glob(f'*{platform}*search*.{ext}')))

        content_files = list(set(content_files))

        print(f"找到 {len(content_files)} 个内容文件")
        for f in content_files:
            print(f"  - {f}")

        video_links = []
        for content_file in content_files:
            try:
                if content_file.suffix == '.json':
                    with open(content_file, 'r', encoding='utf-8') as f:
                        items = json.load(f)
                        if isinstance(items, list):
                            for item in items:
                                extract_video_links(item, video_links, platform)
                        elif isinstance(items, dict):
                            extract_video_links(items, video_links, platform)

                elif content_file.suffix == '.jsonl':
                    with open(content_file, 'r', encoding='utf-8') as f:
                        for line in f:
                            line = line.strip()
                            if line:
                                try:
                                    item = json.loads(line)
                                    extract_video_links(item, video_links, platform)
                                except json.JSONDecodeError:
                                    continue

                elif content_file.suffix == '.csv':
                    import csv
                    with open(content_file, 'r', encoding='utf-8-sig', errors='replace') as f:
                        reader = csv.DictReader(f)
                        print(f"CSV文件列头: {reader.fieldnames}")
                        for row in reader:
                            extract_video_links(row, video_links, platform)

                elif content_file.suffix == '.xlsx' and EXCEL_AVAILABLE:
                    try:
                        workbook = openpyxl.load_workbook(content_file, data_only=True)
                        for sheet_name in workbook.sheetnames:
                            sheet = workbook[sheet_name]
                            headers = None
                            for row_idx, row in enumerate(sheet.iter_rows(values_only=True), 1):
                                if row_idx == 1:
                                    headers = list(row)
                                    continue
                                if not any(row):
                                    continue
                                if headers:
                                    item = dict(zip(headers, row))
                                    extract_video_links(item, video_links, platform)
                        workbook.close()
                        print(f"已处理Excel文件: {content_file}")
                    except Exception as e:
                        print(f"读取Excel文件 {content_file} 时出错: {str(e)}")

            except Exception as e:
                print(f"读取文件 {content_file} 时出错: {str(e)}")

        print(f"从文件中提取到 {len(video_links)} 个视频链接")

        seen_urls = set()
        for idx, (title, url) in enumerate(video_links, 1):
            if url not in seen_urls:
                seen_urls.add(url)
                display_title = title[:50] + '...' if len(title) > 50 else title
                markdown_lines.append(f"{idx}. [{display_title}]({url})")

        if not video_links:
            markdown_lines.append(f"- 暂无视频下载链接")

        markdown_lines.append(f"")

    sanitized_keywords = keywords.replace('/', '_').replace('\\', '_').replace(':', '_').replace('*', '_').replace('?', '_').replace('"', '_').replace('<', '_').replace('>', '_').replace('|', '_')
    markdown_filename = data_dir / f"视频下载链接_{sanitized_keywords}.md"

    with open(markdown_filename, 'w', encoding='utf-8') as f:
        f.write('\n'.join(markdown_lines))

    print(f"\n已生成视频下载链接markdown文件: {markdown_filename}")


def generate_config_json():
    """自动生成 data/config.json 配置文件"""
    data_dir = Path('data')

    platform_templates = {
        'dy': {
            'xlsx_sheets': ["content", "comments"],
            'video_naming_pattern': "dy_{video_id}.mp4"
        },
        'bili': {
            'xlsx_sheets': ["content", "comments"],
            'video_naming_pattern': "bili_{video_id}.mp4"
        },
        'tavily': {
            'xlsx_sheets': ["content"],
            'image_storage_path': "data/images"
        }
    }

    platforms_data = {}

    xlsx_dir = data_dir / 'xlsx'
    if xlsx_dir.exists():
        for xlsx_file in xlsx_dir.glob('*.xlsx'):
            filename = xlsx_file.name
            if filename.startswith('douyin'):
                if 'dy' not in platforms_data:
                    platforms_data['dy'] = {'xlsx': [], 'video_files': []}
                platforms_data['dy']['xlsx'].append(filename)
            elif filename.startswith('bilibili'):
                if 'bili' not in platforms_data:
                    platforms_data['bili'] = {'xlsx': [], 'video_files': []}
                platforms_data['bili']['xlsx'].append(filename)
            elif filename.startswith('tavily'):
                if 'tavily' not in platforms_data:
                    platforms_data['tavily'] = {'xlsx': [], 'image_files': []}
                platforms_data['tavily']['xlsx'].append(filename)

    video_dir = data_dir / 'video'
    if video_dir.exists():
        for video_file in video_dir.glob('*.mp4'):
            filename = video_file.name
            if filename.startswith('dy_'):
                if 'dy' not in platforms_data:
                    platforms_data['dy'] = {'xlsx': [], 'video_files': []}
                if 'video_files' not in platforms_data['dy']:
                    platforms_data['dy']['video_files'] = []
                platforms_data['dy']['video_files'].append(filename)
            elif filename.startswith('bili_'):
                if 'bili' not in platforms_data:
                    platforms_data['bili'] = {'xlsx': [], 'video_files': []}
                if 'video_files' not in platforms_data['bili']:
                    platforms_data['bili']['video_files'] = []
                platforms_data['bili']['video_files'].append(filename)

    images_dir = data_dir / 'images'
    has_images = False
    if images_dir.exists():
        image_files = []
        for ext in ['jpg', 'jpeg', 'png', 'gif', 'webp']:
            image_files.extend([f.name for f in images_dir.glob(f'*.{ext}')])
        if image_files:
            has_images = True
            if 'tavily' not in platforms_data:
                platforms_data['tavily'] = {'xlsx': [], 'image_files': []}
            platforms_data['tavily']['image_files'] = image_files

    platforms_detail = {}
    platforms_list = []

    for platform, data in platforms_data.items():
        template = platform_templates.get(platform, {})
        detail = {}

        if data.get('xlsx'):
            detail['xlsx'] = data['xlsx']
            detail['xlsx_sheets'] = template.get('xlsx_sheets', ["content"])

        if data.get('video_files'):
            detail['video'] = {
                "enabled": True,
                "storage_path": "data/video",
                "naming_pattern": template.get('video_naming_pattern', f"{platform}_{{video_id}}.mp4"),
                "files": data['video_files']
            }

        if data.get('image_files'):
            detail['image'] = {
                "enabled": True,
                "storage_path": "data/images",
                "files": data['image_files']
            }

        if detail:
            platforms_detail[platform] = detail
            platforms_list.append(platform)

    directories = {
        "xlsx": "data/xlsx",
        "video": "data/video",
        "video_config": "data/video/video_config.json"
    }
    
    if has_images:
        directories["images"] = "data/images"
        directories["image_config"] = "data/images/image_config.json"

    config_data = {
        "generated_time": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        "platforms": platforms_list,
        "directories": directories,
        "platforms_detail": platforms_detail
    }

    config_path = data_dir / 'config.json'
    with open(config_path, 'w', encoding='utf-8') as f:
        json.dump(config_data, f, ensure_ascii=False, indent=2)

    print(f"已生成配置文件: {config_path}")
    return config_data


def run_crawler_by_links(platform, video_links, max_comments, enable_sub_comments, enable_video_download):
    """通过视频链接运行爬虫"""
    print(f'=========================================')
    print(f'开始爬取视频评论')
    print(f'平台: {platform}')
    print(f'视频链接数量: {len(video_links.split(";"))}')
    print(f'评论数量: {max_comments}')
    print(f'二级评论: {enable_sub_comments}')
    print(f'视频下载: {enable_video_download}')
    print(f'=========================================')

    platforms = [p.strip() for p in platform.split(',')]

    for p in platforms:
        print(f"\n--- 处理平台: {get_platform_name(p)} ---")
        
        update_config_for_links(p, video_links, max_comments, enable_sub_comments, enable_video_download)

        print("开始爬取指定视频链接...")
        try:
            exit_code = os.system('python main.py')
            if exit_code != 0:
                print(f'警告: 爬虫退出码为 {exit_code}')
        except Exception as e:
            print(f'爬虫执行异常: {str(e)}')

        time.sleep(2)

    try:
        generate_video_download_markdown(platforms, '视频链接')
    except Exception as e:
        print(f'生成markdown异常: {str(e)}')

    try:
        generate_config_json()
    except Exception as e:
        print(f'生成config.json异常: {str(e)}')


def main():
    if len(sys.argv) < 5:
        print('用法: python run_crawler_by_links.py <platform> <video_links> <max_comments> <enable_sub_comments> [enable_video_download]')
        print('示例: python run_crawler_by_links.py "dy" "https://www.douyin.com/video/123456" 20 true true')
        print('平台选项: dy, bili')
        sys.exit(1)

    platform = sys.argv[1]
    video_links = sys.argv[2]
    max_comments = int(sys.argv[3])
    enable_sub_comments = sys.argv[4].lower() == 'true'
    enable_video_download = sys.argv[5].lower() == 'true' if len(sys.argv) > 5 else True

    run_crawler_by_links(platform, video_links, max_comments, enable_sub_comments, enable_video_download)


if __name__ == '__main__':
    main()